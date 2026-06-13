import hashlib

import openpyxl
from pathlib import Path
from openpyxl.styles import Alignment, Font, PatternFill

from ..domain.models import Solucion

COLORES_FIJOS = {
    "111": "aaadad",  # Rojo claro si 111 representa alguna penalización o retén
}


def _write_solution_xlsx_gantt(path: Path, solution: Solucion) -> None:
    # try:
    #     import openpyxl
    #     from openpyxl.styles import Font, PatternFill, Alignment
    # except ImportError:
    #     print("Advertencia: openpyxl no está instalado. No se generará el Excel.")
    #     return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horario_Visual"

    # Caché dinámico para no recalcular colores ni recrear objetos PatternFill
    fills_cacheados = {}

    def obtener_fill(codigo: str) -> PatternFill:
        codigo_norm = codigo.strip()
        if codigo_norm not in fills_cacheados:
            if codigo_norm in COLORES_FIJOS:
                color_hex = COLORES_FIJOS[codigo_norm]
            else:
                color_hex = _generar_color_pastel(codigo_norm)
            fills_cacheados[codigo_norm] = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        return fills_cacheados[codigo_norm]

    # Definir cabeceras
    headers = [
        "ID",
        "Turno",
        "Núcleo",
        "PTD",
        "CON",
        "T_Asignado",
        "Imaginario",
        "Baja_Alta",
        "Slot_Alta",
        "Slot_Baja",
    ]

    turnos = solution.get_turnos()
    controladores = solution.get_controladores()

    if turnos:
        num_slots = len(turnos[0]) // 3
        headers.extend([f"S_{i}" for i in range(num_slots)])

    ws.append(headers)

    # Estilos de cabecera
    fill_header = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    font_header = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    # Construcción de filas y coloreado automático
    for controlador, turno_str in zip(controladores, turnos, strict=True):
        row_data = [
            controlador.id,
            controlador.turno,
            controlador.nucleo,
            controlador.ptd,
            controlador.con,
            controlador.turno_asignado,
            (
                controlador.baja_alta.value
                if hasattr(controlador.baja_alta, "value")
                else str(controlador.baja_alta)
            ),
            controlador.slot_alta,
            controlador.slot_baja,
        ]

        slots_individuales = [turno_str[i : i + 3] for i in range(0, len(turno_str), 3)]
        row_data.extend(slots_individuales)
        ws.append(row_data)

        current_row = ws.max_row

        for idx, slot_val in enumerate(slots_individuales):
            col_excel = 11 + idx
            celda = ws.cell(row=current_row, column=col_excel)

            # Llamamos a la función inteligente que te da el color hasheado o fijo
            celda.fill = obtener_fill(slot_val)
            celda.alignment = Alignment(horizontal="center")

    # Congelar paneles
    ws.freeze_panes = "K2"

    wb.save(path)


def _generar_color_pastel(texto: str) -> str:
    """Genera un color hexadecimal pastel consistente basado en el texto.
    Al ser un hash matemático, 'aao' siempre devolverá el mismo color exacto.
    """
    # Normalizamos el texto
    texto_norm = texto.strip()

    # Creamos un hash a partir del string
    h = hashlib.sha256(texto_norm.encode("utf-8")).hexdigest()

    # Extraemos los canales RGB (Rojo, Verde, Azul)
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)

    # Mezclamos con blanco puro (255) para suavizarlo y hacerlo pastel
    r = (r + 255) // 2
    g = (g + 255) // 2
    b = (b + 255) // 2

    return f"{r:02X}{g:02X}{b:02X}"


def _write_solution_txt(path: Path, solution: Solucion) -> None:
    lines = ["# turnos"]
    lines.extend(solution.get_turnos())
    lines.append("# controladores")
    for controlador in solution.get_controladores():
        lines.append(
            f"id={controlador.id};turno={controlador.turno};nucleo={controlador.nucleo};"
            f"PTD={controlador.ptd};CON={controlador.con};turnoAsignado={controlador.turno_asignado};"
            f"slots_trabajados={controlador.slots_trabajados}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_solution_xlsx(path: Path, solution: Solucion) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos"
    ws.append(["turno_index", "slot_index", "valor"])
    for turno_idx, turno in enumerate(solution.get_turnos()):
        for slot_idx in range(0, len(turno), 3):
            ws.append([turno_idx, slot_idx // 3, turno[slot_idx : slot_idx + 3]])

    ws_ctrl = wb.create_sheet("Controladores")
    ws_ctrl.append(
        [
            "id",
            "turno",
            "nucleo",
            "PTD",
            "CON",
            "turnoAsignado",
            "bajaAlta",
            "slotAlta",
            "slotBaja",
            "slots_trabajados",
        ]
    )
    for controlador in solution.get_controladores():
        ws_ctrl.append(
            [
                controlador.id,
                controlador.turno,
                controlador.nucleo,
                controlador.ptd,
                controlador.con,
                controlador.turno_asignado,
                controlador.baja_alta.value,
                controlador.slot_alta,
                controlador.slot_baja,
                controlador.slots_trabajados,
            ]
        )
    wb.save(path)
